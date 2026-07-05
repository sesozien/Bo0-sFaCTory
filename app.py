import os
import re
import json
import asyncio
import threading
import logging
import streamlit as st
import pandas as pd
from PIL import Image, ImageFilter, ImageDraw

# مكتبات التنسيق الفخم لشيت الإكسيل
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from telegram import Update
    from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, filters, ContextTypes
    HAS_TELEGRAM = True
except ModuleNotFoundError:
    HAS_TELEGRAM = False

import config

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

st.set_page_config(page_title="Bo0sViDClone v12.5 VIP", page_icon="🛸", layout="wide")

# 🔒 التوكن المحفوظ بشكل دائم ومضمون
TOKEN = "8685178390:AAEgzrKz2yHW2oeflsZyZMeSN1Nw0da3vvI" 

EXCEL_PATH = "Amazon_Shopping_Montgk.xlsx"
ACTIVE_LOGO_PATH = "logo.png"
DEFAULT_LOGO_PATH = "default_logo.png"

if not os.path.exists(DEFAULT_LOGO_PATH):
    Image.new('RGBA', (200, 200), color=(255, 75, 75, 255)).save(DEFAULT_LOGO_PATH)
if not os.path.exists(ACTIVE_LOGO_PATH):
    Image.open(DEFAULT_LOGO_PATH).save(ACTIVE_LOGO_PATH)

# دالة التنسيق الملون الفخم المدمجة للبوت
def save_styled_excel(file_path, new_row):
    arabic_headers = ["كود المنتج (SKU)", "اسم وعنوان المنتج", "سعر البيع (جنيه)", "اسم البراند", "الوصف والمميزات", "سعر القطعة جملة", "سعر الشراء الأصلي"]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "كتالوج المنتجات"
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.rightToLeft = True 

    # باليتة الألوان الهادئة الاحترافية (Pastel)
    header_fill = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid") 
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid") 
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Arial", size=10, color="000000")
    thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                         top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

    for col_num, header in enumerate(arabic_headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    rows_to_write = []
    if os.path.exists(file_path):
        try:
            old_df = pd.read_excel(file_path)
            for _, r in old_df.iterrows():
                if pd.notna(r.iloc[0]) and not str(r.iloc[0]).startswith("🛸") and not str(r.iloc[0]).startswith("كود"):
                    rows_to_write.append(list(r[:7]))
        except:
            pass

    rows_to_write.append([
        new_row["sku"], new_row["title"], new_row["standard_price"],
        new_row["brand_name"], new_row["description"], new_row["piece_price"], new_row["original_price"]
    ])

    for row_idx, data in enumerate(rows_to_write, 3):
        current_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.fill = current_fill
            cell.border = thin_border
            
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.0' if isinstance(val, float) else '#,##0'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right" if col_idx in [2, 5] else "center", vertical="center")

    # إمضاء MR Bo0 الفخم بيتحرك تلقائي لآخر الشيت
    sig_row = len(rows_to_write) + 5
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=7)
    sig_cell = ws.cell(row=sig_row, column=1)
    sig_cell.value = "🛸 تم الإنشاء والترتيب بكل فخامة بواسطة: MR Bo0 🛸"
    sig_cell.font = Font(name="Arial", size=12, bold=True, italic=True, color="2A4D69")
    sig_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 28
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(file_path)

if not os.path.exists(EXCEL_PATH):
    save_styled_excel(EXCEL_PATH, {"sku":"-", "title":"-", "standard_price":0, "brand_name":"-", "description":"-", "piece_price":0.0, "original_price":0})

# --- 🎯 محرك الترتيب الذكي اللغوي ---
def advanced_smart_parse(text):
    clean_text = re.sub(r'01[0125]\d{8}', '', text)
    clean_text = clean_text.replace("2026", "").replace("2025", "")
    
    box_count = None
    for kw in config.DOZEN_KEYWORDS:
        pattern = r'(?:' + re.escape(kw) + r'\s*[:\-=\s]*\s*(\d+))|(\d+)\s*' + re.escape(kw)
        match = re.search(pattern, clean_text)
        if match:
            box_count = int(match.group(1) or match.group(2))
            break
    if not box_count or box_count == 0:
        box_count = 12

    total_price = None
    for kw in config.PRICE_KEYWORDS:
        pattern = r'(?:' + re.escape(kw) + r'\s*[:\-=\s]*\s*(\d+))|(\d+)\s*' + re.escape(kw)
        match = re.search(pattern, clean_text)
        if match:
            total_price = int(match.group(1) or match.group(2))
            break

    if "سعر القطعة" in clean_text or "سعر القطعه" in clean_text:
        p_match = re.search(r'(?:سعر القطع[ةه]\s*[:\-=\s]*\s*(\d+))|(\d+)\s*سعر القطع[ةه]', clean_text)
        if p_match:
            piece_val = int(p_match.group(1) or p_match.group(2))
            if total_price == piece_val or not total_price:
                total_price = piece_val * box_count

    if not total_price:
        total_price = 0

    return total_price, box_count

def process_product_image(in_path, out_path, bottom_text="Montgk Brand"):
    img = Image.open(in_path).convert("RGBA")
    w, h = img.size
    
    blurred_img = img.filter(ImageFilter.GaussianBlur(radius=8))
    mask = Image.new("L", (w, h), 0)
    draw = ImageDraw.Draw(mask)
    draw.ellipse((w*0.1, h*0.1, w*0.9, h*0.9), fill=255)
    mask = mask.filter(ImageFilter.GaussianBlur(radius=15))
    img = Image.composite(img, blurred_img, mask)

    if os.path.exists(ACTIVE_LOGO_PATH):
        logo = Image.open(ACTIVE_LOGO_PATH).convert("RGBA")
        logo.thumbnail((int(w*0.25), int(h*0.15)))
        img.paste(logo, (w - logo.size[0] - 15, 15), logo)
        
    draw_obj = ImageDraw.Draw(img)
    try:
        draw_obj.text((20, h - 40), bottom_text, fill=(255, 255, 255, 200))
    except: pass
    img.save(out_path, "PNG")

# ==================== 🛠️ لوحة التحكم الجانبية ====================
with st.sidebar:
    st.markdown("<h2 style='color:#ff4b4b;'>🛸 تحكم مــنتــجـك VIP</h2>", unsafe_allow_html=True)
    st.markdown("### 🖼️ لوحة تغيير اللوجو")
    new_logo = st.file_uploader("ارفع أو الصق لوجو جديد هنا:", type=["png", "jpg", "jpeg"])
    if new_logo is not None:
        Image.open(new_logo).save(ACTIVE_LOGO_PATH)
        st.success("✅ تم تحديث اللوجو بنجاح في السيرفر!")
        
    if st.button("🔄 إزالة اللوجو المخصص"):
        Image.open(DEFAULT_LOGO_PATH).save(ACTIVE_LOGO_PATH)
        st.success("🔄 تم استعادة اللوجو الافتراضي!")
        st.rerun()
        
    st.write("---")
    st.markdown("### 📝 نص البراند السفلي")
    if "brand_text_input" not in st.session_state:
        st.session_state["brand_text_input"] = "Montgk Brand"
    b_text = st.text_input("تعديل النص المطبوع على الصورة:", value=st.session_state["brand_text_input"])
    if b_text != st.session_state["brand_text_input"]:
        st.session_state["brand_text_input"] = b_text
        st.success("📝 تم تعديل الاسم بنجاح!")

# ==================== 🛰️ محرك البوت الذكي (التليجرام) ====================
ASK_SKU, ASK_TITLE, ASK_DESC, ASK_PRICE, ASK_BOX_COUNT, ASK_BRAND = range(6)

async def start_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🛸 مستعد لتلقي صور منتجات مــنتــجـك وتكويدها تلقائياً مع الكينج MR Bo0!")

async def handle_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photo_file = await update.message.photo[-1].get_file()
    raw_path = "bot_raw.png"
    processed_path = "bot_processed.png"
    await photo_file.download_to_drive(raw_path)
    
    current_text = st.session_state.get("brand_text_input", "Montgk Brand")
    process_product_image(raw_path, processed_path, bottom_text=current_text)
    
    await update.message.reply_photo(photo=open(processed_path, 'rb'), caption="⚡ تم معالجة وحقن الصورة باللوجو النشط والاسم الحالي!")
    
    context.user_data['current_product'] = {}
    await update.message.reply_text("📦 **السؤال 1:** اكتب كود المنتج الفريد (Product SKU)؟")
    return ASK_SKU

async def get_sku(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['current_product']['sku'] = update.message.text.strip()
    await update.message.reply_text("🏷️ **السؤال 2:** اسم وعنوان المنتج (Product Title)؟")
    return ASK_TITLE

async def get_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['current_product']['title'] = update.message.text.strip()
    await update.message.reply_text("📝 **السؤال 3:** وصف ومميزات المنتج (Description)؟")
    return ASK_DESC

async def get_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['current_product']['description'] = update.message.text.strip()
    await update.message.reply_text("💰 **السؤال 4:** كام سعر الشراء / الجملة الإجمالي؟")
    return ASK_PRICE

async def get_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['current_product']['original_price'] = int(update.message.text.strip())
        await update.message.reply_text("📦 **السؤال 5:** عدد القطع داخل البوكس أو الدستة؟")
        return ASK_BOX_COUNT
    except:
        await update.message.reply_text("⚠️ اكتب رقم صحيح للسعر:")
        return ASK_PRICE

async def get_box_count(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data['current_product']['box_count'] = int(update.message.text.strip())
        await update.message.reply_text("🏭 **السؤال 6:** اسم البراند؟ (أو اكتب /skip لو عام)")
        return ASK_BRAND
    except:
        await update.message.reply_text("⚠️ اكتب رقم صحيح:")
        return ASK_BOX_COUNT

async def get_brand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['current_product']['brand'] = update.message.text.strip()
    return await save_bot_data(update, context)

async def skip_brand(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['current_product']['brand'] = "Generic"
    return await save_bot_data(update, context)

async def save_bot_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    prod = context.user_data['current_product']
    selling_price = int(prod['original_price'] * 1.25)
    piece_price = round(selling_price / prod['box_count'], 1)
    
    new_row = {
        "sku": prod['sku'], "title": prod['title'], "standard_price": selling_price,
        "brand_name": prod['brand'], "description": prod['description'], "piece_price": piece_price, "original_price": prod['original_price']
    }
    
    save_styled_excel(EXCEL_PATH, new_row)
    
    commercial_post = (
        f"📦 **{prod['title']}**\n🔢 كود المنتج: `{prod['sku']}`\n\n"
        f"💬 {prod['description']}\n\n"
        f"🔥 سعر العرض الخاص من مــنتــجـك: {selling_price} ج!\n"
        f"📌 (سعر القطعة جملة واصل عليك بـ {piece_price} ج بس! 💣)\n\n"
        f"#MR_Bo0\n#Montgk-منتجك"
    )
    await update.message.reply_text(commercial_post)
    return ConversationHandler.END

def run_bot_loop():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    app = Application.builder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[MessageHandler(filters.PHOTO, handle_image)],
        states={
            ASK_SKU: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_sku)],
            ASK_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_title)],
            ASK_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_desc)],
            ASK_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_price)],
            ASK_BOX_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_box_count)],
            ASK_BRAND: [CommandHandler("skip", skip_brand), MessageHandler(filters.TEXT & ~filters.COMMAND, get_brand)],
        },
        fallbacks=[]
    )
    app.add_handler(CommandHandler("start", start_bot))
    app.run_polling()

if HAS_TELEGRAM and "bot_thread_started" not in st.session_state and TOKEN != "":
    st.session_state["bot_thread_started"] = True
    threading.Thread(target=run_bot_loop, daemon=True).start()

# ==================== 💻 واجهة العرض الرئيسية ====================
st.subheader("📊 كتالوج وجدول منتجات أمازون والمنصات الموحد")

test_input = st.text_area("أدخل نص كشط المنتج لتجربة الفحص الذكي والترتيب اللغوي:")
if test_input:
    parsed_price, parsed_box = advanced_smart_parse(test_input)
    st.write(f"🔍 السعر الملقوط: **{parsed_price} ج** | 📦 عدد القطع: **{parsed_box}**")
    chosen_orig_price = st.number_input("💵 السعر الأصلي للمنتج:", min_value=0, max_value=1000000, value=int(parsed_price))

if os.path.exists(EXCEL_PATH):
    try:
        df_display = pd.read_excel(EXCEL_PATH)
        df_clean = df_display[df_display.iloc[:, 0].astype(str).str.contains("🛸") == False]
        st.dataframe(df_clean, use_container_width=True)
    except:
        st.info("📊 الجدول يتم تهيئته الآن واستقبال البيانات الفخمة...")

st.markdown("<br><p style='text-align: center; color: #2a4d69; font-weight: bold;'>🛸 تم التطوير والأداء بكل فخامة بواسطة: MR Bo0 🛸</p>", unsafe_allow_html=True)
